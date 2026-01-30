"""Extract formulas from Excel file to understand priority queue calculations"""
import openpyxl

# Load the Excel file
excel_path = "/Users/danielahurtadolange/Dropbox/Documents/Work and studies/Kellogg/Teaching/2026 Winter/Queue spreadsheet stuff/Queue.xlsx"

try:
    wb = openpyxl.load_workbook(excel_path, data_only=False)
    wb_data = openpyxl.load_workbook(excel_path, data_only=True)

    # Focus on the Infinite Queue sheet where priorities are
    ws = wb["Infinite Queue"]
    ws_data = wb_data["Infinite Queue"]

    print("=== KEY CELL REFERENCES ===\n")

    # Check what H5, H7, and H11 are
    key_cells = ['H5', 'H7', 'H11', 'H14']
    for cell_ref in key_cells:
        cell = ws[cell_ref]
        cell_data = ws_data[cell_ref]
        print(f"{cell_ref}:")
        if cell.data_type == 'f':
            print(f"  Formula: {cell.value}")
        else:
            print(f"  Value: {cell.value}")
        print(f"  Calculated: {cell_data.value}")
        print()

    # Also check the labels in column G
    print("\n=== COLUMN G LABELS (H column values) ===")
    for row_idx in range(1, 15):
        g_cell = ws.cell(row=row_idx, column=7)  # Column G
        h_cell = ws.cell(row=row_idx, column=8)  # Column H
        h_cell_data = ws_data.cell(row=row_idx, column=8)
        if g_cell.value:
            formula_str = ""
            if h_cell.data_type == 'f':
                formula_str = f" [Formula: {h_cell.value}]"
            print(f"Row {row_idx}: {g_cell.value} = {h_cell_data.value}{formula_str}")

    print("\n\n=== PRIORITY FORMULAS DECODED ===\n")

    # Now let's decode the formulas
    print("Based on the formulas found:")
    print("\nFor priority class k (k=1 is highest priority):")
    print("\n  Intermediate calculation F(k):")
    print("    F(1) = 1 - fraction(1) * ρ")
    print("    F(k) = F(k-1) - fraction(k) * ρ   for k > 1")
    print("\n  Ii(k) - Average number in queue for class k:")
    print("    Ii(1) = Ii_total * fraction(1) * (1 - ρ) / F(1)")
    print("    Ii(k) = Ii_total * fraction(k) * (1 - ρ) / (F(k) * F(k-1))   for k > 1")
    print("\n  Ti(k) - Average wait time for class k:")
    print("    Ti(k) = Ii(k) / (fraction(k) * λ)")
    print("\nWhere:")
    print("  - λ (H5) = arrival rate")
    print("  - Ii_total (H7) = overall average number in queue (from M/M/c)")
    print("  - ρ (H11) = utilization")

except Exception as e:
    print(f"Error reading Excel file: {e}")
    import traceback
    traceback.print_exc()
